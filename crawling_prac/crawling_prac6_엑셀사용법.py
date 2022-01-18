import openpyxl

wb = openpyxl.load_workbook('')
ws = wb['오징어게임']

ws['A3'] = 456
ws['B3'] = '성기훈'

wb.save()